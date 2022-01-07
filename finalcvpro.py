import cv2, sys, numpy, os
#for pop up message
from tkinter import messagebox as tkMessageBox
from tkinter import *
#for excel readi & write
import openpyxl
from openpyxl import Workbook
#for date
import datetime
#from datetime import datetime as dt
from datetime import date
#for image
import tkinter as tk
from PIL import Image, ImageTk
from PIL import Image

TK_SILENCE_DEPRECATION=1

#LARGE_FONT= ("Verdana", 20)
LARGE_FONT= ("Times New Roman", 20)
'''
checktime=datetime.datetime.now()
fo = open("checkfile.txt", "wb")

fo.write( "System is now logged out at %s"%str(checktime));
# Close opend file
fo.close()
'''
class SeaofBTCapp(tk.Tk):

    def __init__(self, *args, **kwargs):
        
        tk.Tk.__init__(self, *args, **kwargs)
        container = tk.Frame(self)

        container.pack(side="top", fill="both", expand = True)

        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}

        for F in (StartPage, PageOne, PageTwo, Login, sign, report):

            frame = F(container, self)

            self.frames[F] = frame

            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(StartPage)

       
    def show_frame(self, cont):

        frame = self.frames[cont]
        frame.tkraise()
    
        self.menubar = Menu(self)
	 
class StartPage(tk.Frame):

    def __init__(self, parent, controller):
       	self.controller = controller  
        tk.Frame.__init__(self,parent)
        label = tk.Label(self, text="Start Page", font=LARGE_FONT)
        label.pack(pady=10,padx=10)

	#college logo image
        load = Image.open("UTA_A-logo.jpg")
        render = ImageTk.PhotoImage(load)
        # labels can be text or images
        img = tk.Label(self, image=render)
        img.image = render
        img.place(x=450, y=250)

        #attendance management system image
        load = Image.open("Screen Shot 2017-02-27 at 11.24.39 AM.png")
        render = ImageTk.PhotoImage(load)
        # labels can be text or images
        img = tk.Label(self, image=render)
        img.image = render
        img.place(x=200, y=100)

        button = tk.Button(self, text="Start",font=LARGE_FONT,
                            command=lambda: controller.show_frame(Login))
        button.pack()
        button.place(x=500, y=600)

        button2 = tk.Button(self, text="Exit",font=LARGE_FONT,
                            command=lambda: controller.show_frame(PageTwo))
        button2.pack()
        button2.place(x=850, y=600)

class Login(tk.Frame):

    def __init__(self, parent, controller):
        self.controller = controller
        tk.Frame.__init__(self, parent)
        label = tk.Label(self, text="Admin Login!!!", font=LARGE_FONT)
        label.pack(pady=10,padx=10)

        label1=tk.Label(self, text='Enter username and Password', font=LARGE_FONT).pack(side=TOP)
        label2=tk.Label(self, text="Username ", font=LARGE_FONT).pack(side=TOP)
        ent = tk.Entry(self)
        ent.pack(side=TOP)
        label3=tk.Label(self, text="Password ", font=LARGE_FONT).pack(side=TOP)
        ent2 = tk.Entry(self, show="*")
        ent2.pack(side=TOP)
        btn = tk.Button(self, text="Submit", font=LARGE_FONT, command=lambda: self.reply(ent.get(),ent2.get()))
        btn.pack(side=TOP)
        btn2 = tk.Button(self, text="Cancel", font=LARGE_FONT, command=lambda: controller.show_frame(PageTwo) )
        btn2.pack(side=TOP)

    def reply(self, name, passw):
                        if name =='darshan':
                                if passw =='hero':
                                        tkMessageBox.showinfo(title='Reply', message='Hello %s!' % name)
                                        self.controller.show_frame(PageOne)
                                else:
                                        tkMessageBox.showinfo(title='Reply', message=' %s! is a invalid pin try again' % name)
                        else:
                                tkMessageBox.showinfo(title='Reply', message= 'Invalid username or password try again')


class PageOne(tk.Frame):
	def __init__(self, parent, controller):
		self.controller = controller
		tk.Frame.__init__(self, parent)
		label = tk.Label(self, text="Main Menu!!!", font=LARGE_FONT)
		label.pack(pady=10,padx=10)
		name=None	
		#for logout 
		btnlog = tk.Button(self, text="Logout ", font=LARGE_FONT, command=lambda: controller.show_frame(PageTwo))
		btnlog.pack()
		btnlog.place(x=1300,y=0)
		
		#For Recognition
		label1 = tk.Label(self, text="To start attendance marking press Start!", font=LARGE_FONT)
		label1.place(x=600,y=100)
		btnrec = tk.Button(self, text="Start ", font=LARGE_FONT, command=lambda: self.recognize(name))
		btnrec.pack()
		btnrec.place(x=700,y=150)

		#For New Entry
		label2 = tk.Label(self, text="For new entry press Sign up!", font=LARGE_FONT)
		label2.place(x=300,y=300)
		btnnew = tk.Button(self, text="Sign Up", font=LARGE_FONT,command=lambda: controller.show_frame(sign))
		btnnew.pack()
		btnnew.place(x=400,y=350)
		#For Reset
		label3 = tk.Label(self, text="To reset the Attendance sheet press Reset!", font=LARGE_FONT)
		label3.place(x=600,y=500)
		btnnre = tk.Button(self, text="Reset", font=LARGE_FONT,command=lambda: self.reset(name))
		btnnre.pack()
		btnnre.place(x=700,y=550)

		#For Report
		label4 = tk.Label(self, text="For Report menu press Reports!", font=LARGE_FONT)
		label4.place(x=900,y=300)
		btnnrep = tk.Button(self, text="Reports", font=LARGE_FONT,command=lambda: controller.show_frame(report))
		btnnrep.pack()
		btnnrep.place(x=1000,y=350)

	def recognize(self,name):
		top = Tk()
		top.title('Margin Time')
		top.iconbitmap('py-blue-trans-out.ico')

		Label(top, text="Enter the allowable minutes:").pack(side=TOP)
		ent = Entry(top)
		ent.pack(side=TOP)
		btn = Button(top, text="Submit", command=(lambda: reply(ent.get())))
		btn.pack(side=LEFT)

		#top.mainloop()
		def reply(margin):		
			#initialization of variables
			min=datetime.datetime.now().minute
			margin=int(margin)
			min=(min+margin)%60
			coll=0
			flag=0
			flag1=0
			todo=None
			j=0

			#excel loading Existing file
			book = openpyxl.load_workbook('Attendance.xlsx')
			sheet = book.active

			#initializing date and week day
			for i in range(1, sheet.max_column+2):
				if sheet.max_column+1==i:
					today = datetime.datetime.now().date()
					todo=today
					day=today.strftime("%A")
					sheet.cell(row=1, column=i).value=todo
					sheet.cell(row=2, column=i).value=day
					coll=i

			#initially marking everyone as Absent
			for i in range(3, sheet.max_row+1):
				sheet.cell(row=i, column=coll).value='A'
			book.save('Attendance.xlsx')

			#main recognition program
			size = 4
			haar_file = 'haarcascade_frontalface_default.xml'
			datasets = 'datasets'

			# Part 1: Create fisherRecognizer
			print('Training...')
			tkMessageBox.showinfo(title='working', message=' I am Learning please wait! ')
			# Create a list of images and a list of corresponding names
			(images, labels, names, id) = ([], [], {}, 0)

			for (subdirs, dirs, files) in os.walk(datasets):
				for subdir in dirs:
					names[id] = subdir
					subjectpath = os.path.join(datasets, subdir)
					for filename in os.listdir(subjectpath):
						path = subjectpath + '/' + filename
						label = id
						images.append(cv2.imread(path, 0))
						labels.append(int(label))
					id += 1
			(width, height) = (130, 100)

			# Create a Numpy array from the two lists above
			(images, labels) = [numpy.array(lis) for lis in [images, labels]]

			# OpenCV trains a model from the images
			# NOTE FOR OpenCV2: remove '.face'
			model = cv2.face.FisherFaceRecognizer_create()
			model.train(images, labels)

			if flag==0:
				tkMessageBox.showinfo(title='Start', message=' Now I am ready! ')
				flag=1

			# Part 2: Use fisherRecognizer on camera stream
			face_cascade = cv2.CascadeClassifier(haar_file)
			webcam = cv2.VideoCapture(0)
			while True:
				(_, im) = webcam.read()
				gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
				faces = face_cascade.detectMultiScale(gray, 1.3, 5)
				for (x,y,w,h) in faces:
					cv2.rectangle(im,(x,y),(x+w,y+h),(255,0,0),2)
					face = gray[y:y + h, x:x + w]
					face_resize = cv2.resize(face, (width, height))
					#Try to recognize the face
					prediction = model.predict(face_resize)
					cv2.rectangle(im, (x, y), (x + w, y + h), (0, 255, 0), 3)

					if prediction[1]<500:
						cv2.putText(im,'%s' % (names[prediction[0]]),(x-10, y-10), cv2.FONT_HERSHEY_PLAIN,1,(0, 255, 0))
						str=names[prediction[0]]

						for i in range(3, sheet.max_row+1):
							str1=sheet.cell(row=i, column=2).value
							if str1==str :
								j=j+1
								if j==10:
									j=0
									sheetval=sheet.cell(row=i, column=coll).value
									print(sheetval)
									if sheetval!='P' and sheetval!='L':

										book = openpyxl.load_workbook('Attendance.xlsx')
										sheet = book.active
										temp=datetime.datetime.now().minute
										#print temp
										if min>=int(temp):
											sheet.cell(row=i, column=coll).value='P'
										else:
											sheet.cell(row=i, column=coll).value='L'
										book.save('Attendance.xlsx')

										tym = datetime.datetime.now().time()
										book = openpyxl.load_workbook('Attendance.xlsx')
										sheetB = book.active
										sheetB.cell(row=i, column=coll).value=tym
										book.save('time.xlsx')		
										tkMessageBox.showinfo(title='Start', message=' your Attendance is marked!')
	
					else:
						cv2.putText(im,'not recognized',(x-10, y-10), cv2.FONT_HERSHEY_PLAIN,1,(0, 255, 0))

				cv2.imshow('OpenCV', im)
				key = cv2.waitKey(10)
				if key == 27:
					break    

	def reset(self,name):
		datasets='datasets'
		wb = Workbook()
		ws = wb.active
		ws['A1'] = 'Sr No.'
		ws['B1'] = 'Names'
		rw=3
		for (subdirs, dirs, files) in os.walk(datasets):
			for subdir in dirs:
				ws.cell(row=rw, column=1).value=rw-2
				ws.cell(row=rw, column=2).value=subdir
				rw+=1
		wb.save('Attendance.xlsx')

		book = openpyxl.load_workbook('Attendance.xlsx')
		sheet = book.active
		wb.save('time.xlsx')
		tkMessageBox.showinfo(title='Done!', message='Both the file are reset! ')

class sign(tk.Frame):

	def __init__(self, parent, controller):
		self.controller = controller
		tk.Frame.__init__(self, parent)
		label = tk.Label(self, text="Sign Up!!!", font=LARGE_FONT)
		label.pack(pady=10,padx=10)


		#For new entry
		label1=tk.Label(self, text='For new entry:',font=LARGE_FONT).pack(side=TOP)
		label2=tk.Label(self, text="Name ",font=LARGE_FONT).pack(side=TOP)
		ent = tk.Entry(self)
		ent.pack(side=TOP)
        
		label3=tk.Label(self, text="Roll no. ",font=LARGE_FONT).pack(side=TOP)

		ent2 = tk.Entry(self)
		ent2.pack(side=TOP)
		str=ent.get()+ent2.get() 
		print(str)
		btn = tk.Button(self, text="Submit", font=LARGE_FONT, command=lambda: self.dataset(ent.get(),ent2.get()))
		btn.pack(side=TOP)

		button2 = tk.Button(self, text="Exit",font=LARGE_FONT,
                            command=lambda: controller.show_frame(PageTwo))
		button2.pack(side=TOP)

	def dataset(self, name, roll):
		x='_'
		name=name+x
		name=name+roll
		#appending attendance
		wb = openpyxl.load_workbook('Attendance.xlsx')
		ws = wb['Sheet']		
		r=ws.max_row+1
		ws.cell(row=r, column=1).value=r-1
		ws.cell(row=r, column=2).value=name
		#ws.cell(row=r, column=3).value=roll
		wb.save('Attendance.xlsx')

		wb = openpyxl.load_workbook('Attendance.xlsx')
		ws = wb['Sheet']		
		r=ws.max_row+1
		ws.cell(row=r, column=1).value=r-1
		ws.cell(row=r, column=2).value=name
		#ws.cell(row=r, column=3).value=roll
		wb.save('Time.xlsx')

		haar_file = 'haarcascade_frontalface_default.xml'
		datasets = 'datasets'  #All the faces data will be present this folder
		sub_data = name     #These are sub data sets of folder, for my faces I've used my name



		path = os.path.join(datasets, sub_data)
		if not os.path.isdir(path):
			os.mkdir(path)
			(width, height) = (130, 100)    # defining the size of images


			face_cascade = cv2.CascadeClassifier(haar_file)
			webcam = cv2.VideoCapture(0) #'0' is use for my webcam, if you've any other camera attached use '1' like this

			# The program loops until it has 100 images of the face.
			count = 0
			while count < 31:
				(_, im) = webcam.read()
				gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
				faces = face_cascade.detectMultiScale(gray, 1.3, 4)
				for (x,y,w,h) in faces:
					cv2.rectangle(im,(x,y),(x+w,y+h),(255,0,0),2)
					face = gray[y:y + h, x:x + w]
					face_resize = cv2.resize(face, (width, height))
					cv2.imwrite('%s/%s.png' % (path,count), face_resize)
				count += 1
		
				cv2.imshow('OpenCV', im)
				key = cv2.waitKey(10)
				if key == 27:
					break

class report(tk.Frame):

    def __init__(self, parent, controller):
        self.controller = controller
        tk.Frame.__init__(self, parent)
        label = tk.Label(self, text="Report Menu!!!", font=LARGE_FONT)
        label.pack(pady=10,padx=10)
	
	#for call by value
        dar=None

	#for logout
        btnlog = tk.Button(self, text="Logout ", font=LARGE_FONT, command=lambda: controller.show_frame(PageTwo))
        btnlog.pack()
        btnlog.place(x=1300,y=0)

	#For Recognition
        label1 = tk.Label(self, text="For single person report click Individual button!", font=LARGE_FONT)
        label1.place(x=550,y=100)
        btnrec = tk.Button(self, text="Individual ", font=LARGE_FONT, command=lambda: self.individual(dar))
        btnrec.pack()
        btnrec.place(x=700,y=150)

        #For New Entry
        label2 = tk.Label(self, text="For date wise report press DateWise button!", font=LARGE_FONT)
        label2.place(x=150,y=350)
        btnnew = tk.Button(self, text="Datewise", font=LARGE_FONT,command=lambda: self.datewise(dar))
        btnnew.pack()
        btnnew.place(x=300,y=400)

        #For Reset
        label3 = tk.Label(self, text="For weekly report press Weekly button!", font=LARGE_FONT)
        label3.place(x=550,y=600)
        btnnre = tk.Button(self, text="Weekly", font=LARGE_FONT,command=lambda: self.weekly(dar))
        btnnre.pack()
        btnnre.place(x=700,y=650)

        #For Report
        label4 = tk.Label(self, text="For monthly report press Monthly button!", font=LARGE_FONT)
        label4.place(x=950,y=350)
        btnnrep = tk.Button(self, text="Monthly", font=LARGE_FONT,command=lambda: self.monthly(dar))
        btnnrep.pack()
        btnnrep.place(x=1050,y=400)
	
	#For Report
        label4 = tk.Label(self, text="For overall report press Overall button!", font=LARGE_FONT)
        label4.place(x=550,y=350)
        btnnrep = tk.Button(self, text="Overall", font=LARGE_FONT,command=lambda: self.overall(dar))
        btnnrep.pack()
        btnnrep.place(x=700,y=400)

    def individual(self,dar):
	
        top = Tk()
        top.title('Name')
        top.iconbitmap('py-blue-trans-out.ico')

        Label(top, text="Enter the name in form(name_rollno)").pack(side=TOP)
        ent = Entry(top)
        ent.pack(side=TOP)
        btn = Button(top, text="Submit", command=(lambda: reply(ent.get())))
        btn.pack(side=LEFT)

        book = openpyxl.load_workbook('Attendance.xlsx')
        sheet = book.active

        def reply(name):
                name2=None
                for i in range(3, sheet.max_row+1):

                        sheet_name=sheet.cell(row=i, column=2).value
                        #print name
                        #print sheet_name
                        if sheet_name==name:
                                name2=sheet_name
                                break
                if name2==sheet_name:
                        #print 'success'
                        report(name)
                else:   
                        tkMessageBox.showinfo(title='Error', message=' %s! is not in the list try again' % name)                   
        def report(name):
                total=0
                count_p=0
                count_l=0
                wb = Workbook()
                sheet2 = wb.active
                for i in range(3, sheet.max_row+1):
                        sheet_name=sheet.cell(row=i, column=2).value
                        if sheet_name==name:
                                sheet2['A1']='Date'
                                sheet2['A2']=name

                                name_row=i
                                for j in range(3, sheet.max_column+1):
                                        date=sheet.cell(row=1, column=j+2).value
                                        sheet2.cell(row=1,column=j+1).value=date
                                        attendance=sheet.cell(row=name_row, column=j+2).value
                                        sheet2.cell(row=3,column=j+1).value=attendance
                                        total += 1
                                        if attendance == 'P' or attendance == 'L' :
                                                count_p += 1
                                        if attendance == 'L':
                                                count_l +=1

                                percentage = 100 * float(count_p)/float(total)
                                break
                sheet2['A5']='Total no. of classes taken :'
                sheet2['D5']=total
                sheet2['A6']='Total no. of classes attended :'
                sheet2['D6']=count_p
                sheet2['A7']='Total no. of late mark :'
                sheet2['D7']=count_l
                sheet2['A8']='Attendance percentage :'
                sheet2['D8']=float(percentage)
                wb.save('%s_Report.xlsx'%sheet_name)
                tkMessageBox.showinfo(title='Success', message='Report generated in a Excel file named Report.xlsx')

	
    def datewise(self,dar):
        wb = openpyxl.load_workbook('Workbook.xlsx')
        sheet = wb['Sheet1']

        col=sheet.max_column
        c1=(sheet.cell(row=1, column=3).value).date()
        sh=(sheet.cell(row=1, column=col).value).date()

        top = Tk()
        top.title('Date')
        top.iconbitmap('py-blue-trans-out.ico')
        Label(top, text='The report is available from '+str(c1)+' to '+str(sh)).pack(side=TOP)
        Label(top, text="Enter the start date (in the form yyyy-mm-dd) ").pack(side=TOP)
        ent = Entry(top)
        ent.pack(side=TOP)
        Label(top, text="Enter the end date (in the form yyyy-mm-dd) ").pack(side=TOP)
        ent2 = Entry(top)
        ent2.pack(side=TOP)
        btn = Button(top, text="Submit", command=(lambda: reply(ent.get(),ent2.get())))
        btn.pack(side=LEFT)
    	
    def reply(start_date,end_date):
        count_p=0
        count_l=0
        total=0
        start_col=0
        end_col=0
        x=0
        y=0

        wb = Workbook()
        sheet2 = wb.active

        for j in range(3, sheet.max_column+1):
                temp=sheet.cell(row=1,column=j).value.date()
                if start_date==str(temp):
                        start_col=j
                        x=1
                if end_date==str(temp):
                        end_col=j
                        y=1
                if x!=1:
                    tkMessageBox.showinfo(title='Start_date Error!', message=' start_date does not matched try again ')		
                if y!=1:
                    tkMessageBox.showinfo(title='End_date Error!', message=' end_date does not matched try again ')

                for i in range(1, sheet.max_row+1):
                    for j in range(1, sheet.max_column+1):
                        if j<3:
                            sheet2.cell(row=i,column=j).value=sheet.cell(row=i,column=j).value
                        if j==start_col:
                            diff=j-3
                for i in range(1, sheet.max_row+1):
                   for j in range(start_col, end_col+1):
                       sheet2.cell(row=i,column=j-diff).value=str(sheet.cell(row=i,column=j).value)

                max_col=sheet2.max_column
                sheet2.cell(row=1,column=max_col+1).value='Total classes taken'
                sheet2.cell(row=1,column=max_col+2).value='Total classes attended'
                sheet2.cell(row=1,column=max_col+3).value='Total no. of late mark'
                sheet2.cell(row=1,column=max_col+4).value='Attendance Percentage'

                for i in range(2, sheet2.max_row+1):
                    for j in range(start_col,end_col+1):
                       attendance=sheet.cell(row=i, column=j).value
                       total += 1
                       if attendance == 'P' or attendance == 'L':
                           count_p += 1
                       if attendance == 'L':
                           count_l +=1
                    sheet2.cell(row=i,column=max_col+1).value=total
                    sheet2.cell(row=i,column=max_col+2).value=count_p
                    sheet2.cell(row=i,column=max_col+3).value=count_l
                    percentage = 100 * float(count_p)/float(total)
                    sheet2.cell(row=i,column=max_col+4).value=float(percentage)
                    total=0
                    count_p=0
                    count_l=0
                    percentage=0
                #print "Report generated"
                wb.save('Datewise_Attendance_Report_from_%s_to.xlsx'%start_date)
                tkMessageBox.showinfo(title='Success', message='Report generated in a Excel file named Datewise_Attendance_Report.xlsx')
    
    def weekly(self,dar):
        vermo=0

    
        wb = openpyxl.load_workbook('dummy.xlsx')
        sheet = wb['Sheet1']

        col=sheet.max_column

        for i in range(4, sheet.max_column+1):
                if 'Monday'==str(sheet.cell(row=2,column=i).value):
                        start_date=(sheet.cell(row=1,column=i).value).date()
                        #print sheet.cell(row=2,column=i).value
                        #print start_date
                        vermo +=1
                if 'Saturday'==str(sheet.cell(row=2,column=i).value):
                        end_date=(sheet.cell(row=1,column=i).value).date()
                        #print end_date
                        self.replyw(start_date,end_date,vermo)
                if i==sheet.max_column:
                        end_date=(sheet.cell(row=1,column=i-1).value).date()
                        self.replyw(start_date,end_date,vermo)
    def replyw(self,start_date,end_date,vermo):
                count_p=0
                count_l=0
                total=0
                start_col=0
                end_col=0
                x=0
                y=0

                wb = openpyxl.load_workbook('dummy.xlsx')
                sheet = wb['Sheet1']			
                wb = Workbook()
                sheet2 = wb.active

                for j in range(4, sheet.max_column):
                     temp=(sheet.cell(row=1,column=j).value)
                     #print temp.date()
                     if start_date==temp.date():
                        start_col=j
                        x=1
                     if end_date==temp.date():
                        end_col=j
                        y=1
                if x!=1:
                    tkMessageBox.showinfo(title='Start_date Error!', message=' start_date does not matched try again ')		
                if y!=1:
                    tkMessageBox.showinfo(title='End_date Error!', message=' end_date does not matched try again ')

                for i in range(1, sheet.max_row+1):
                    for j in range(1, sheet.max_column+1):
                        if j<3:
                            sheet2.cell(row=i,column=j).value=sheet.cell(row=i,column=j).value
                        if j==start_col:
                            diff=j-4
                for i in range(1, sheet.max_row+1):
                     for j in range(start_col, end_col+1):
                         sheet2.cell(row=i,column=j-diff).value=str(sheet.cell(row=i,column=j).value)
		
                max_col=sheet2.max_column
                sheet2.cell(row=1,column=max_col+1).value='Total classes taken'
                sheet2.cell(row=1,column=max_col+2).value='Total classes attended'
                sheet2.cell(row=1,column=max_col+3).value='Total no. of late mark'
                sheet2.cell(row=1,column=max_col+4).value='Attendance Percentage'

                for i in range(3, sheet2.max_row+1):
                    for j in range(start_col,end_col+1):
                        attendance=sheet.cell(row=i, column=j).value
                        total += 1
                        if attendance == 'P' or attendance == 'L':
                            count_p += 1
                        if attendance == 'L':
                            count_l +=1
                    sheet2.cell(row=i,column=max_col+1).value=total
                    sheet2.cell(row=i,column=max_col+2).value=count_p
                    sheet2.cell(row=i,column=max_col+3).value=count_l
                    percentage = 100 * float(count_p)/float(total)
                    sheet2.cell(row=i,column=max_col+4).value=float(percentage)
               	    total=0
                    count_p=0
                    count_l=0
                    percentage=0
                #print "Report generated"
                #vermo+=1
                wb.save('Weekly_report_week_%d_to.xlsx'%vermo)
                tkMessageBox.showinfo(title='Success', message='Report generated in a Excel file named Weekly_eport%d.xlsx'%vermo)

    def monthly(self,dar):
        wb = openpyxl.load_workbook('Workbook.xlsx')
        sheet = wb['Sheet1']

        col=sheet.max_column
        c1=(sheet.cell(row=1, column=3).value).date()
        sh=(sheet.cell(row=1, column=col).value).date()

        top = Tk()
        top.title('Month')
        top.iconbitmap('py-blue-trans-out.ico')
        Label(top, text='The report is available from '+str(c1)+' to '+str(sh)).pack(side=TOP)
        Label(top, text="Enter the month in form (for feb enter 2) ").pack(side=TOP)
        ent = Entry(top)
        ent.pack(side=TOP)
        btn = Button(top, text="Submit", command=(lambda: reply(ent.get())))
        btn.pack(side=LEFT)
	
        def reply(mnth):
            moth=int(mnth)	
            count_p=0
            count_l=0
            total=0
            start_col=0
            end_col=0
            x=0
            y=0
            first=0
            second=0

            #for extracting start date of the month and the end date			


            wb = Workbook()
            sheet2 = wb.active

            for j in range(3, sheet.max_column+1):
                temp=sheet.cell(row=1,column=j).value.date()
                temp2=sheet.cell(row=1,column=j).value.month
                if moth==temp2 and first==0:
                    tart_date=temp
                    start_col=j
                    x=1
                    first=1
                if (moth+1)%12 == temp2 or j == sheet.max_column:
                    if second==0:
                       end_date=temp
                       end_col=j
                       y=1
                       second==1
            if x!=1 or y!=1:
                tkMessageBox.showinfo(title='Erroe!', message=' Entered Month does not matched try again ')		
		
            for i in range(1, sheet.max_row+1):
                for j in range(1, sheet.max_column+1):
                    if j<3:
                        sheet2.cell(row=i,column=j).value=sheet.cell(row=i,column=j).value
                    if j==start_col:
                        diff=j-3
            for i in range(1, sheet.max_row+1):
                for j in range(start_col, end_col+1):
                    sheet2.cell(row=i,column=j-diff).value=str(sheet.cell(row=i,column=j).value)
		
            max_col=sheet2.max_column
            sheet2.cell(row=1,column=max_col+1).value='Total classes taken'
            sheet2.cell(row=1,column=max_col+2).value='Total classes attended'
            sheet2.cell(row=1,column=max_col+3).value='Total no. of late mark'
            sheet2.cell(row=1,column=max_col+4).value='Attendance Percentage'
		
            for i in range(2, sheet2.max_row+1):
                for j in range(start_col,end_col+1):
                    attendance=sheet.cell(row=i, column=j).value
                    total += 1
                    if attendance == 'P' or attendance == 'L':
                        count_p += 1
                    if attendance == 'L':
                        count_l +=1
                		
                sheet2.cell(row=i,column=max_col+1).value=total
                sheet2.cell(row=i,column=max_col+2).value=count_p
                sheet2.cell(row=i,column=max_col+3).value=count_l
                percentage = 100 * float(count_p)/float(total)
                sheet2.cell(row=i,column=max_col+4).value=float(percentage)
                total=0
                count_p=0
                count_l=0
                percentage=0
            #print "Report generated"
            wb.save('Month_%d_Attendance report.xlsx'%moth)
            tkMessageBox.showinfo(title='Success', message='Report generated in a Excel file named Month_Attendance report.xlsx')
	

    def overall(self,dar):
        wb = openpyxl.load_workbook('Workbook.xlsx')
        sheet = wb['Sheet1']

        col=sheet.max_column
        c1=(sheet.cell(row=1, column=3).value).date()
        sh=(sheet.cell(row=1, column=col).value).date()
        self.reply(c1,sh)
	    	
    def reply(self,start_date,end_date):
        count_p=0
        count_l=0
        total=0
        start_col=0
        end_col=0
        x=0
        y=0
		
        wb = openpyxl.load_workbook('Workbook.xlsx')
        sheet = wb['Sheet1']	

        wb = Workbook()
        sheet2 = wb.active

        for j in range(3, sheet.max_column+1):
            temp=sheet.cell(row=1,column=j).value.date()
            if start_date==temp:
                start_col=j
                x=1
            if end_date==temp:
                end_col=j
                y=1
        if x!=1:
            tkMessageBox.showinfo(title='Start_date Error!', message=' start_date does not matched try again ')		
        if y!=1:
            tkMessageBox.showinfo(title='End_date Error!', message=' end_date does not matched try again ')

        for i in range(1, sheet.max_row+1):
            for j in range(1, sheet.max_column+1):
                if j<3:
                    sheet2.cell(row=i,column=j).value=sheet.cell(row=i,column=j).value
                if j==start_col:
                    diff=j-3
                for i in range(1, sheet.max_row+1):
                        for j in range(start_col, end_col+1):
                             sheet2.cell(row=i,column=j-diff).value=str(sheet.cell(row=i,column=j).value)

                max_col=sheet2.max_column
                sheet2.cell(row=1,column=max_col+1).value='Total classes taken'
                sheet2.cell(row=1,column=max_col+2).value='Total classes attended'
                sheet2.cell(row=1,column=max_col+3).value='Total no. of late mark'
                sheet2.cell(row=1,column=max_col+4).value='Attendance Percentage'

                for i in range(2, sheet2.max_row+1):
                    for j in range(start_col,end_col+1):
                        attendance=sheet.cell(row=i, column=j).value
                        total += 1
                        if attendance == 'P' or attendance == 'L':
                            count_p += 1
                        if attendance == 'L':
                            count_l +=1
                        sheet2.cell(row=i,column=max_col+1).value=total
                        sheet2.cell(row=i,column=max_col+2).value=count_p
                        sheet2.cell(row=i,column=max_col+3).value=count_l
                        percentage = 100 * float(count_p)/float(total)
                        sheet2.cell(row=i,column=max_col+4).value=float(percentage)
                        total=0
                        count_p=0
                        count_l=0
                        percentage=0
                #print "Report generated"
                wb.save('Overall_Attendance_Report_from_%s_to.xlsx'%start_date)
                tkMessageBox.showinfo(title='Success', message='Report generated in a Excel file named Overall_Attendance_Report.xlsx')

	
		
class PageTwo(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        label = tk.Label(self, text="Back to Home or Exit!!!", font=LARGE_FONT)
        label.pack(pady=10,padx=10)

        button1 = tk.Button(self, text="Back to Home",
                            command=lambda: controller.show_frame(StartPage))
        button1.pack()

        button2 = tk.Button(self, text="Exit",
                            command=self.check)
        button2.pack()
	
    def check(self):
        try:
            os.remove("checkfile.txt")
        except OSError:
            pass
        self.quit()
        


app = SeaofBTCapp()
app.attributes('-fullscreen', True)
wid = app.winfo_screenwidth()
app.mainloop()

